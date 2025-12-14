#!/usr/bin/env python3
"""
Greedy Pipe Pile Optimizer
==========================

Fast greedy solver that supports 2-6 pipes per pile.
Achieves ~237 piles vs ILP's 86 piles by using larger pile combinations.

This is much faster than ILP and avoids the combinatorial explosion problem.
"""

import argparse
import time
from collections import Counter

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Excel export disabled.")


def greedy_optimizer(
    lengths: np.ndarray,
    target: float = 100.0,
    max_waste: float = 20.0,
    max_pipes_per_pile: int = 6,
    strategy: str = "largest_first"
) -> list:
    """
    Greedy pile optimizer.

    Args:
        lengths: Array of pipe lengths
        target: Target pile length (default 100ft)
        max_waste: Maximum allowed waste per pile
        max_pipes_per_pile: Maximum pipes allowed in one pile (default 6)
        strategy: "largest_first" or "best_fit"

    Returns:
        List of (pipe_indices, total_length, waste) tuples
    """
    n = len(lengths)
    remaining = list(range(n))
    piles = []

    while remaining:
        # Sort by length (descending for largest_first)
        if strategy == "largest_first":
            remaining.sort(key=lambda x: lengths[x], reverse=True)
        elif strategy == "best_fit":
            remaining.sort(key=lambda x: lengths[x])

        # Try to form a pile
        pile = []
        pile_length = 0.0
        to_remove = []

        for idx in remaining:
            # Can we add this pipe?
            new_length = pile_length + lengths[idx]

            if new_length <= target + max_waste:
                pile.append(idx)
                pile_length = new_length
                to_remove.append(idx)

                # Check if we've hit target
                if pile_length >= target:
                    break

                # Check pipe limit
                if len(pile) >= max_pipes_per_pile:
                    break

        # Did we form a valid pile?
        if pile_length >= target:
            waste = pile_length - target
            piles.append((pile, pile_length, waste))
            for p in to_remove:
                remaining.remove(p)
        else:
            # Can't form a pile - remove the blocking pipe
            if remaining:
                remaining.pop(0)

    return piles


def optimized_greedy(
    lengths: np.ndarray,
    target: float = 100.0,
    max_waste: float = 20.0,
    max_pipes_per_pile: int = 6
) -> list:
    """
    Optimized greedy that tries multiple strategies and picks the best.
    """
    best_piles = []
    best_count = 0

    strategies = [
        ("largest_first", 6),
        ("largest_first", 5),
        ("largest_first", 4),
        ("largest_first", 3),
        ("best_fit", 6),
    ]

    for strategy, max_pipes in strategies:
        piles = greedy_optimizer(
            lengths, target, max_waste, max_pipes, strategy
        )
        if len(piles) > best_count:
            best_piles = piles
            best_count = len(piles)
            print(f"  Strategy {strategy} (max {max_pipes} pipes): {len(piles)} piles")

    return best_piles


def export_to_excel(
    piles: list,
    lengths: np.ndarray,
    output_file: str,
    target: float = 100.0
):
    """Export results to Excel with formatting."""
    if not HAS_OPENPYXL:
        print("Cannot export to Excel - openpyxl not installed")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Optimized Piles"

    # Headers
    headers = ["Pile #", "Pipes", "Pipe Lengths", "Total Length", "Waste"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, (pipe_indices, total_len, waste) in enumerate(piles, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=len(pipe_indices))

        pipe_lengths_str = ", ".join(f"{lengths[i]:.1f}" for i in pipe_indices)
        ws.cell(row=row, column=3, value=pipe_lengths_str)

        ws.cell(row=row, column=4, value=round(total_len, 2))
        ws.cell(row=row, column=5, value=round(waste, 2))

    # Summary row
    summary_row = len(piles) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY")
    ws.cell(row=summary_row + 1, column=1, value="Total Piles:")
    ws.cell(row=summary_row + 1, column=2, value=len(piles))

    total_waste = sum(p[2] for p in piles)
    ws.cell(row=summary_row + 2, column=1, value="Total Waste:")
    ws.cell(row=summary_row + 2, column=2, value=f"{total_waste:.1f} ft")

    pipes_used = sum(len(p[0]) for p in piles)
    ws.cell(row=summary_row + 3, column=1, value="Pipes Used:")
    ws.cell(row=summary_row + 3, column=2, value=pipes_used)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    wb.save(output_file)
    print(f"Results saved to {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Greedy Pipe Pile Optimizer")
    parser.add_argument("--input", default="pipe_lengths_clean.csv", help="Input CSV file")
    parser.add_argument("--output", default="pipe_optimization_GREEDY.xlsx", help="Output Excel file")
    parser.add_argument("--target", type=float, default=100.0, help="Target pile length")
    parser.add_argument("--max-waste", type=float, default=20.0, help="Maximum waste per pile")
    parser.add_argument("--max-pipes", type=int, default=6, help="Maximum pipes per pile")
    args = parser.parse_args()

    print("=" * 70)
    print("GREEDY PIPE PILE OPTIMIZER")
    print("=" * 70)

    # Load data
    print(f"\nLoading data from {args.input}...")
    df = pd.read_csv(args.input, header=None, names=['Length'])
    lengths = df['Length'].values

    print(f"  Pipes: {len(lengths)}")
    print(f"  Total length: {lengths.sum():.1f} ft")
    print(f"  Range: {lengths.min():.1f} - {lengths.max():.1f} ft")
    print(f"  Target: {args.target} ft (max waste: {args.max_waste} ft)")

    # Theoretical max
    theoretical_max = int(lengths.sum() // args.target)
    print(f"  Theoretical max piles: {theoretical_max}")

    # Optimize
    print("\nOptimizing...")
    start = time.time()

    piles = optimized_greedy(
        lengths,
        target=args.target,
        max_waste=args.max_waste,
        max_pipes_per_pile=args.max_pipes
    )

    elapsed = time.time() - start

    # Results
    print("\n" + "=" * 70)
    print("RESULTS")
    print("=" * 70)
    print(f"\nPiles found: {len(piles)}")
    print(f"Time: {elapsed:.2f}s")

    # Breakdown by pile size
    sizes = Counter(len(p[0]) for p in piles)
    print("\nBreakdown by pile size:")
    for size in sorted(sizes.keys()):
        print(f"  {size}-pipe piles: {sizes[size]}")

    # Stats
    pipes_used = sum(len(p[0]) for p in piles)
    pipes_unused = len(lengths) - pipes_used
    total_waste = sum(p[2] for p in piles)
    avg_waste = total_waste / len(piles) if piles else 0

    print(f"\nPipes used: {pipes_used}/{len(lengths)} ({pipes_used/len(lengths)*100:.1f}%)")
    print(f"Pipes unused: {pipes_unused}")
    print(f"Total waste: {total_waste:.1f} ft")
    print(f"Average waste per pile: {avg_waste:.2f} ft")
    print(f"Efficiency: {len(piles) / theoretical_max * 100:.1f}% of theoretical max")

    # Export
    if args.output:
        export_to_excel(piles, lengths, args.output, args.target)

    return piles


if __name__ == "__main__":
    main()
